import os
import pandas as pd
from garminconnect import Garmin
from datetime import datetime


# Läs in Garmin Connect-uppgifter från inställningar.txt
import configparser
import os
settings_path = os.path.join(os.path.dirname(__file__), 'inställningar.txt')

# Läs in inställningar
GARMIN_USER = None
GARMIN_PASS = None
DAYS_TO_FILL = 7
if os.path.exists(settings_path):
    with open(settings_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('GARMIN_USER='):
                GARMIN_USER = line.strip().split('=', 1)[1]
            elif line.startswith('GARMIN_PASS='):
                GARMIN_PASS = line.strip().split('=', 1)[1]
            elif line.startswith('DAYS_TO_FILL='):
                try:
                    DAYS_TO_FILL = int(line.strip().split('=', 1)[1])
                except Exception:
                    DAYS_TO_FILL = 7
if not GARMIN_USER or not GARMIN_PASS:
    raise Exception('Fyll i GARMIN_USER och GARMIN_PASS i inställningar.txt')

EXCEL_FILE = "garmin_traningspass.xlsx"

# Logga in på Garmin Connect
api = Garmin(GARMIN_USER, GARMIN_PASS)
api.login()

from datetime import timedelta, date
# Hämta aktiviteter från de senaste DAYS_TO_FILL dagarna
today = date.today()
days_ago = today - timedelta(days=DAYS_TO_FILL)
all_activities = []
index = 0
batch_size = 20
while True:
    batch = api.get_activities(index, batch_size)
    if not batch:
        break
    filtered = [act for act in batch if pd.to_datetime(act["startTimeLocal"]).date() >= days_ago]
    all_activities.extend(filtered)
    # Avbryt om sista i batch är äldre än DAYS_TO_FILL dagar
    if any(pd.to_datetime(act["startTimeLocal"]).date() < days_ago for act in batch):
        break
    index += batch_size
activities = all_activities

# Skapa DataFrame av aktiviteterna
new_df = pd.DataFrame([
    {
        "activityId": act["activityId"],
        "date": act["startTimeLocal"],
        "type": act["activityType"]["typeKey"],
        "distance_km": act.get("distance", 0) / 1000,
        "duration_min": act.get("duration", 0) / 60,
        "calories": act.get("calories", 0),
        "name": act.get("activityName", ""),
        "description": act.get("description", ""),
        "maxHR": act.get("maxHR", ""),
        "averageHR": act.get("averageHR", ""),
        "hrTimeInZone_1": act.get("hrTimeInZone_1", ""),
        "hrTimeInZone_2": act.get("hrTimeInZone_2", ""),
        "hrTimeInZone_3": act.get("hrTimeInZone_3", ""),
        "hrTimeInZone_4": act.get("hrTimeInZone_4", ""),
        "hrTimeInZone_5": act.get("hrTimeInZone_5", "")
    }
    for act in activities
])

# Om filen finns, läs in och slå ihop utan dubbletter


# Om filen finns, uppdatera rader där datum, duration och namn matchar, annars lägg till nytt pass
if os.path.exists(EXCEL_FILE):
    print(f"Läser in befintlig fil {EXCEL_FILE} för att uppdatera...")
    old_df = pd.read_excel(EXCEL_FILE)
    # Se till att alla kolumner från new_df finns i old_df
    for col in new_df.columns:
        if col not in old_df.columns:
            old_df[col] = ""
    # Se till att kolumnordningen är samma
    old_df = old_df[new_df.columns]
    # Skapa en kopia för att uppdatera
    updated_df = old_df.copy()
    for _, new_row in new_df.iterrows():
        mask = (
            (updated_df["date"] == new_row["date"]) &
            (updated_df["duration_min"] == new_row["duration_min"]) &
            (updated_df["name"] == new_row["name"])
        )
        if mask.any():
            updated_df.loc[mask, :] = new_row.values
        else:
            updated_df = pd.concat([updated_df, pd.DataFrame([new_row])], ignore_index=True)
    combined = updated_df
else:
    print(f"Skapar ny fil {EXCEL_FILE}")
    combined = new_df


# Säkerställ att alla kolumner finns och i rätt ordning
cols = ["activityId", "date", "type", "distance_km", "duration_min", "calories", "name", "description"]
for col in cols:
    if col not in combined.columns:
        combined[col] = ""
combined = combined[cols]

combined.to_excel(EXCEL_FILE, index=False)

# Skriv ut beskrivning för varje pass
for i, row in new_df.iterrows():
    print(f"Pass {i+1} - {row['date']} - {row['name']}")
    print(f"Typ: {row['type']}, Distans: {row['distance_km']} km, Tid: {row['duration_min']:.1f} min, Kalorier: {row['calories']}")
    print(f"Beskrivning: {row['description']}")
    # Summera pulszoner till låg, medel, hög
    låg = sum([float(row.get(f"hrTimeInZone_{z}", 0) or 0) for z in range(1, 3)])
    medel = float(row.get("hrTimeInZone_3", 0) or 0)
    hög = sum([float(row.get(f"hrTimeInZone_{z}", 0) or 0) for z in range(4, 6)])

    #hög = float(row.get("hrTimeInZone_5", 0) or 0)
    def sek_to_hms(sek):
        sek = int(sek)
        h = sek // 3600
        m = (sek % 3600) // 60
        s = sek % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    print(f"Låg puls: {sek_to_hms(låg)}")
    print(f"Medel puls: {sek_to_hms(medel)}")
    print(f"Hög puls: {sek_to_hms(hög)}")
    print("-")


# Skriv activityId från senaste aktivitet till cell A1 i annan Excel-fil

from openpyxl import load_workbook
MAKRO_EXCEL = None
if os.path.exists(settings_path):
    with open(settings_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('MAKRO_EXCEL='):
                MAKRO_EXCEL = line.strip().split('=', 1)[1]
if not MAKRO_EXCEL:
    print('FEL: Skriv in korrekt filnamn för xlsm i inställningar.txt, t.ex. MAKRO_EXCEL=2024-2025 dagbok test.xlsm')
    exit(1)
makro_excel = MAKRO_EXCEL
if os.path.exists(makro_excel):
    wb = load_workbook(makro_excel, keep_vba=True)
    ws = wb.active
    from openpyxl.styles import numbers
    # Fyll vilopuls för varje dag, även vilodagar
    for offset in range(0, DAYS_TO_FILL+1):
        dag = days_ago + timedelta(days=offset)
        ytd_int = dag.timetuple().tm_yday
        excel_row = ytd_int * 4 + 65 * 4 + 1
        try:
            hr_data = api.get_heart_rates(dag.strftime("%Y-%m-%d"))
            resting_hr = hr_data.get("restingHeartRate", "")
        except Exception:
            resting_hr = ""
        if ws[f"Y{excel_row}"].value in [None, ""]:
            ws[f"Y{excel_row}"] = resting_hr
    # Fyll aktiviteter
    ytd_count = {}
    for i, row in new_df.iterrows():
        try:
            dt = pd.to_datetime(row["date"])
            ytd_int = dt.dayofyear
            if ytd_int not in ytd_count:
                ytd_count[ytd_int] = 0
            # Kontrollera distans
            try:
                dist = float(row["distance_km"])
            except Exception:
                dist = 0
            if ytd_count[ytd_int] < 4:
                excel_row = ytd_int * 4 + 65 * 4 + 1 + ytd_count[ytd_int]
                if ws[f"F{excel_row}"].value in [None, ""]:
                    ws[f"F{excel_row}"] = row["name"]
                # Tid till kolumn G
                total_seconds = int(float(row["duration_min"]) * 60)
                if ws[f"G{excel_row}"].value in [None, ""]:
                    ws[f"G{excel_row}"] = total_seconds / 86400
                # Skriv '1' i kolumn I
                if ws[f"I{excel_row}"].value in [None, ""]:
                    ws[f"I{excel_row}"] = 1
                # Skriv låg, medel, hög puls till J, K, L
                def sek_to_hms(sek):
                    sek = int(sek)
                    h = sek // 3600
                    m = (sek % 3600) // 60
                    s = sek % 60
                    return f"{h:02d}:{m:02d}:{s:02d}"
                låg = sum([float(row.get(f"hrTimeInZone_{z}", 0) or 0) for z in range(1, 4)])
                medel = float(row.get("hrTimeInZone_4", 0) or 0)
                hög = float(row.get("hrTimeInZone_5", 0) or 0)
                if ws[f"J{excel_row}"].value in [None, ""]:
                    ws[f"J{excel_row}"] = låg / 86400
                if ws[f"K{excel_row}"].value in [None, ""]:
                    ws[f"K{excel_row}"] = medel / 86400
                if ws[f"L{excel_row}"] .value in [None, ""]:
                    ws[f"L{excel_row}"] = hög / 86400
                # Skriv distans med max 2 decimaler i kolumn M, men bara om >= 0,5 km
                if dist >= 0.5:
                    dist_str = f"{dist:.2f}".replace(".", ",")
                else:
                    dist_str = ""
                if ws[f"M{excel_row}"].value in [None, ""]:
                    ws[f"M{excel_row}"] = dist_str
                # Skriv maxpuls till kolumn P och medelpuls till kolumn Q
                max_hr = row["maxHR"] if "maxHR" in row and pd.notnull(row["maxHR"]) else ""
                avg_hr = row["averageHR"] if "averageHR" in row and pd.notnull(row["averageHR"]) else ""
                print(f"DEBUG: max_hr={max_hr}, avg_hr={avg_hr}, excel_row={excel_row}")
                if ws[f"P{excel_row}"].value in [None, ""]:
                    ws[f"P{excel_row}"] = max_hr
                if ws[f"Q{excel_row}"].value in [None, ""]:
                    ws[f"Q{excel_row}"] = avg_hr

                # Skriv beskrivning till kolumn X på rad (ytd_int * 4 + 65 * 4 + 1) + 1
                desc_row = ytd_int * 4 + 65 * 4 + 2
                prev_desc = ws[f"X{desc_row}"].value if ws[f"X{desc_row}"].value else ""
                new_desc = row["description"]
                if new_desc and ws[f"X{desc_row}"].value in [None, ""]:
                    ws[f"X{desc_row}"] = str(new_desc)
                print(f"Skrev namn '{row['name']}' till F{excel_row}, tid till G{excel_row}, '1' till L{excel_row}, distans '{dist_str}' till M{excel_row}, maxpuls '{max_hr}' till P{excel_row}, medelpuls '{avg_hr}' till Q{excel_row}, beskrivning till X{desc_row} i {makro_excel}")
                ytd_count[ytd_int] += 1
            else:
                print(f"Hoppar över aktivitet {row['activityId']} på dag {ytd_int}, max 4 per dag.")
        except Exception as e:
            print(f"Kunde inte beräkna YTD för aktivitet {row['activityId']}: {e}")
    wb.save(makro_excel)

print(f"Uppdaterade {EXCEL_FILE} med de senaste 5 passen.")

